# Copyright 2024, Polytechnique Montreal and contributors
# This file is licensed under the MIT License.
# License text available at https://opensource.org/licenses/MIT

# Note: This script includes functions that generate the locales libelles files.
# These functions are intended to be invoked from the generate_survey.py script.
import os  # For interacting with the operating system
from glob import glob, escape  # For file path matching
import ruamel.yaml  # For working with YAML files
import openpyxl  # For reading Excel files
from helpers.generator_helpers import (
    get_headers,
    sheet_exists
)

# Initialize YAML parser
yaml = ruamel.yaml.YAML()
yaml.indent(sequence=4, offset=4, mapping=4)
yaml.width = 80


# Class for handling various text formatting notations
class ValueReplacer:
    # Various HTML and markdown notations
    startBoldHtml = "<strong>"
    endBoldHtml = "</strong>"
    boldNotation = "**"

    startOblique = '<span class="_pale _oblique">'
    endOblique = "</span>"
    obliqueNotation = "__"

    startGreen = '<span style="color: green;">'
    endGreen = "</span>"
    greenNotation = "_green_"

    startRed = '<span style="color: red;">'
    endRed = "</span>"
    redNotation = "_red_"

    # Static methods for replacing notations with proper HTML tags
    @staticmethod
    def replaceStartEnd(string, notation, startReplaced, endReplaced):
        # Replaces notations with corresponding start/end tags in the string
        replacedStr = string
        if notation in replacedStr and replacedStr.count(notation) % 2 == 0:
            replacedCount = 0
            while notation in replacedStr:
                replaceWith = startReplaced if replacedCount % 2 == 0 else endReplaced
                replacedStr = replacedStr.replace(notation, replaceWith, 1)
                replacedCount += 1
        return replacedStr

    # Main replace function applying all notations
    @staticmethod
    def replace(string):
        # Replaces newlines with <br> tags and applies other notations
        replacedStr = string.replace("\n", "<br />")
        # replaced each bold, oblique, green and red notations by proper tags
        replacedStr = ValueReplacer.replaceStartEnd(
            replacedStr,
            ValueReplacer.boldNotation,
            ValueReplacer.startBoldHtml,
            ValueReplacer.endBoldHtml,
        )
        replacedStr = ValueReplacer.replaceStartEnd(
            replacedStr,
            ValueReplacer.obliqueNotation,
            ValueReplacer.startOblique,
            ValueReplacer.endOblique,
        )
        replacedStr = ValueReplacer.replaceStartEnd(
            replacedStr,
            ValueReplacer.greenNotation,
            ValueReplacer.startGreen,
            ValueReplacer.endGreen,
        )
        replacedStr = ValueReplacer.replaceStartEnd(
            replacedStr,
            ValueReplacer.redNotation,
            ValueReplacer.startRed,
            ValueReplacer.endRed,
        )
        return replacedStr


# Class for managing translations in a specific language and section
class TranslationLangNs:
    def __init__(self, excel_file_path):
        self.modified = False
        self.data = {}
        self.file = excel_file_path
        self.startBoldHtml = "<strong>"
        self.endBoldHtml = "</strong>"

    # Convert string to YAML format
    def stringToYaml(self, str):
        if "\n" in str:
            return ruamel.yaml.scalarstring.FoldedScalarString(str)
        if len(str) > 76:
            return ruamel.yaml.scalarstring.FoldedScalarString(str)
        return str

    # Load existing translations from the YAML file
    def loadCurrentTranslations(self):
        with open(self.file, mode="r") as stream:
            try:
                translationData = yaml.load(stream)
                self.data = {}
                for path in translationData:
                    self.data[path] = self.stringToYaml(translationData[path])
            except Exception as err:
                print(f"Error loading yaml file {err}")
                raise Exception("Error loading translation yaml file " + self.file)

    # Save modifications back to the YAML file
    def save(self):
        if self.modified:
            with open(self.file, "w", encoding="utf-8") as file:
                file.write(
                    "# This file was automatically generated by the Evolution Generator.\n"
                )
                file.write(
                    "# The Evolution Generator is used to automate the creation of consistent, reliable code.\n"
                )
                file.write("# Any changes made to this file will be overwritten.\n\n")
                yaml.dump(self.data, file)
            print(f"Generate {self.file.replace('\\', '/')} successfully")

    # Add a new translation or update an existing one
    def addTranslation(self, path, value, overwrite, keepMarkdown):
        if not overwrite and path in self.data:
            return

        value = value.replace("[nom]", r"{{nickname}}")

        # Replace with HTML tags
        if not keepMarkdown:
            value = ValueReplacer.replace(value)

        self.data[path] = self.stringToYaml(value)
        self.modified = True


# Class for managing translations for all languages and sections
class TranslationData:
    def __init__(self, libelles_output_folder_path):
        self.translations = {}  # Dictionary to store translations
        self.libelles_output_folder_path = libelles_output_folder_path  # Path to the locales directory

    # Add translations for a specific language and section
    def addTranslations(self, lang, section, translations):
        if not lang in self.translations:
            self.translations[lang] = {}
        self.translations[lang][section] = translations

    # Save all translations to their respective files
    def save(self):
        for lang in self.translations:
            for section in self.translations[lang]:
                self.translations[lang][section].save()

    # Add a new translation to the specified language, section, and path
    def addTranslation(self, lang, section, path, value, overwrite, keepMarkdown):
        try:
            if not lang in self.translations:
                self.translations[lang] = {}
            if not section in self.translations[lang]:
                self.translations[lang][section] = TranslationLangNs(
                    os.path.join(self.libelles_output_folder_path, lang, section + ".yml")
                )
            self.translations[lang][section].addTranslation(
                path, value, overwrite, keepMarkdown
            )
        except Exception as e:
            print(f"Exception occurred for {lang} {section} {path}: {e}")
            raise e


# Class for managing the overall translation process
class FillLocalesTranslations:
    def __init__(self, excel_file_path, libelles_output_folder_path, overwrite, section):
        self.excel_file_path = excel_file_path
        self.libelles_output_folder_path = libelles_output_folder_path
        self.overwrite = overwrite
        self.section = section
        self.allTranslations = TranslationData(libelles_output_folder_path)
        super().__init__()

    # Load existing translations from YAML files
    def loadCurrentTranslations(self):
        ymlFiles = glob(escape(self.libelles_output_folder_path) + "/**/*.yml")
        for translationFile in ymlFiles:
            path = os.path.normpath(os.path.dirname(translationFile))
            paths = path.split(os.sep)
            lang = paths[len(paths) - 1]
            section = os.path.splitext(os.path.basename(translationFile))[0]
            translationNs = TranslationLangNs(translationFile)
            translationNs.loadCurrentTranslations()
            self.allTranslations.addTranslations(lang, section, translationNs)

    # Save all translations back to YAML files
    def saveAllTranslations(self):
        self.allTranslations.save()

    def __getTranslationsFromLocaleSheet(self):
        try:
            
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
            sheet_exists(workbook, "Labels")
            sheet = workbook["Labels"]  # Get Widgets sheet

            if len(list(sheet.rows)) <= 1:
                raise Exception(f"Not enough rows in the Labels sheet")
            
            # Add the locales from the locales sheet in a dictionary to be added when the key is met
            # Get headers from the first row
            headers = get_headers(
                sheet,
                expected_headers=["key"],
                sheet_name="Labels",
            )

            locales_dictionary = {}
            for row in list(sheet.rows)[1:]:
                    # Create a dictionary from the row values and headers
                    current_locale = dict(zip(headers, (cell.value for cell in row)))
                    locales_dictionary[current_locale['key']] = current_locale
            return locales_dictionary

        except Exception as e:
            print(f"Exception occurred in get locales sheet: {e}")
            return {}

    def __addTranslationsFromLocales(self, section, question_name, row):
        try:
            # Code copied from the fillLocales.py script 
            keepMarkdown = row['md'] if 'md' in row else 0
            if keepMarkdown == '1':
                keepMarkdown = True
            else:
                keepMarkdown = False
            for key in row:
                # Skip known columns
                if key == 'namespace' or key == 'key' or key == 'md':
                    continue
                # Do not process empty string
                if row[key] == '' or row[key] is None:
                    continue
                lngContext = key.split('_', 1)
                # If language part does not have 2 characters, the column should not be processed
                if len(lngContext[0]) > 2:
                    continue
                translationKey = question_name
                if len(lngContext) > 1:
                    translationKey += '_' + lngContext[1]
                self.allTranslations.addTranslation(lngContext[0], section, translationKey, row[key], self.overwrite, keepMarkdown)

        except Exception as e:
            print(f"Exception occurred in add Translations from locale: {e}")
            return {}

    # Function to add translations from Excel input file to the translations data
    def addTranslationsFromExcel(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
            sheet = workbook["Widgets"]  # Get Widgets sheet

            locales_dictionary = self.__getTranslationsFromLocaleSheet()

            # Parse the widget sheet to add the translations
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # TODO: Take row value with header name instead of index
                question_name = row[0]
                section = row[3]
                path = row[5]
                fr = row[6]
                en = row[7]
                # keepMarkdown = row[11]
                keepMarkdown = False

                if question_name in locales_dictionary:
                    self.__addTranslationsFromLocales(section, question_name, locales_dictionary[question_name])
                else:
                    if fr is not None:
                        self.allTranslations.addTranslation(
                            "fr", section, path, fr, self.overwrite, keepMarkdown
                        )
                    if en is not None:
                        self.allTranslations.addTranslation(
                            "en", section, path, en, self.overwrite, keepMarkdown
                        )
                

        except Exception as e:
            print(f"Exception occurred in addTranslationsFromExcel: {e}")
            raise e


# Function to generate the libelles locales files
def generate_libelles(
    excel_file_path, libelles_output_folder_path, overwrite=False, section=None
):
    try:
        # Initialize the FillLocalesTranslations task with provided parameters
        task = FillLocalesTranslations(
            excel_file_path, libelles_output_folder_path, overwrite, section
        )
        task.loadCurrentTranslations()
        task.addTranslationsFromExcel()
        task.saveAllTranslations()
    except Exception as e:
        print(f"An error occurred: {e}")
        raise e
